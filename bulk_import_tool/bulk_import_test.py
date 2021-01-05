import unittest
from datetime import datetime
from bulk_import_tool import ImportTools
import openpyxl
import pyodbc


class BulkImportTest(unittest.TestCase):
    
    def setUp(self):
        self.impt = ImportTools()
        file_name = 'IMM import template - test.xlsx'
        self.impt._get_file(file_name)
        self.impt._get_prog_info()
        self.impt.discipline = 'inv'
        self.impt.area_cd = 'natural'
        self.maxDiff = None
    
    def test_find_persons(self):
        value = self.impt._find_person_organization(type="Person")
        test_values = {'Hugh MacIntosh': [6755], 'Evan Harley': [6168],
                       'Meg Sugrue': [16213], 'David Stewart': ['NEW?'],
                       'Henry Choong': [2767, 4659], 'Heidi Gartner': [2430, 4829]}
        self.assertEqual(test_values, value)


    def test_find_organization(self):
        value = self.impt._find_person_organization(type="Organization")
        test_values = {'Washington Department of Fish and Wildlife': [130]}
        self.assertEqual(test_values, value)

    def test_find_relevant_column(self):
        test_methods = ['Person', 'Taxon', 'Sites', 'Events']
        value = []
        test_values = [22, 75, 76, 117]
        test_values.extend([i for i in range(6, 73)])
        test_values = list(set(test_values))
        for method in test_methods:
            value.extend(self.impt._find_relevant_column(method))
        self.assertEqual(sorted(test_values), sorted(value))

    def test_split_persons(self):
        value = self.impt._split_persons('Evan Harley, Hugh MacIntosh; Meg Sugrue| Dave Stewart: test')
        self.assertEqual(['Evan Harley', 'Hugh MacIntosh', 'Meg Sugrue', 'Dave Stewart', 'test'], value)

    def test_find_taxa(self):
        value = self.impt._find_taxa()
        test_values = {'Cancer productus': [98672], 'Cancer magister': ['NEW?']}  # Cancer magister: 98675
        self.assertEqual(test_values, value)

    def test_query_taxa(self):
        test_taxa = ['Tryphon', 'Heterocladium macounii Best', 'Trivias', 'Homoglaea hircina', 'Acanthodiaptomus']
        value = []
        test_values = [70215058, 90223816, 'NEW?', 70212504, 70219886, 70219887, 70219888,
                       70219889, 70219890, 70219898, 85836, 80047655]
        for taxon in test_taxa:
            value.extend(self.impt._query_taxa(taxon))
        self.assertEqual(test_values, value)
   
    def test_generate_site(self):
        value = self.impt._generate_sites()
        test_values = {'VS101579':
                       {"Collector's Site ID": 'VS101579',
                        'Elevation (max)': 15,
                        'Elevation (min)': 14,
                        'Elevation note': None,
                        'Elevation unit': 'm',
                        'Biogeoclimatic': None,
                        'Biozone': None,
                        'Continent': 'North America',
                        'Country': 'Canada',
                        'County': None,
                        'District': None,
                        'Ecoprovince': None,
                        'Fossil Ref. No.': None,
                        'Mine Name': None,
                        'Natural Region': None,
                        'Park/ER/IR': None,
                        'Province/State': 'British Columbia',
                        'Range/Township/Section': None,
                        'Water Body': None,
                        'Description':	'Fannin tower, 2nd floor',
                        'Discipline': None,
                        'Location Name': 'Victoria: RBCM Collections building',
                        'Reference': None,
                        'Remarks': None,
                        'Notes (Date)': None,
                        'Notes (Note)': None,
                        'Notes (Title)': None,
                        'Accuracy': None,
                        'Approximate': None,
                        'Latitude': 48.419603,
                        'Latitude Stop': None,
                        'Longitude': -123.3706457,
                        'Longitude Stop': None,
                        'N.A. Datapoint': None,
                        'Non-NTS Map Reference': None,
                        'NTS Map Reference': None,
                        'UTM Datapoint': None,
                        'UTM Easting': None,
                        'UTM Northing': None,
                        'UTM Zone':	None,
                        'Primary drainage':	None,
                        'Secondary drainage': None,
                        'Tertiary drainage': None,
                        }, 'VS101580':
                       {"Collector's Site ID": 'VS101580',
                        'Elevation (max)': 5,
                        'Elevation (min)': 4,
                        'Elevation note': None,
                        'Elevation unit': 'm',
                        'Biogeoclimatic': None,
                        'Biozone': None,
                        'Continent': 'North America',
                        'Country': 'Canada',
                        'County': None,
                        'District': None,
                        'Ecoprovince': None,
                        'Fossil Ref. No.': None,
                        'Mine Name': None,
                        'Natural Region': None,
                        'Park/ER/IR': None,
                        'Province/State': 'British Columbia',
                        'Range/Township/Section': None,
                        'Water Body': None,
                        'Description':	'Clifford Carl Hall',
                        'Discipline': None,
                        'Location Name': 'Victoria: RBCM Exhibits building',
                        'Reference': None,
                        'Remarks': None,
                        'Notes (Date)': None,
                        'Notes (Note)': None,
                        'Notes (Title)': None,
                        'Accuracy': None,
                        'Approximate': None,
                        'Latitude': 48.419957,
                        'Latitude Stop': None,
                        'Longitude': -123.3688604,
                        'Longitude Stop': None,
                        'N.A. Datapoint': None,
                        'Non-NTS Map Reference': None,
                        'NTS Map Reference': None,
                        'UTM Datapoint': None,
                        'UTM Easting': None,
                        'UTM Northing': None,
                        'UTM Zone':	None,
                        'Primary drainage':	None,
                        'Secondary drainage': None,
                        'Tertiary drainage': None,
                        }}

        self.assertEqual(test_values, value)
    
    def test_generate_sites_write_site_id(self):
        test_values = ['VS101579', 'VS101580']
        self.impt._generate_sites()
        values = [self.impt.ws.cell(row=i, column=49).value for i in range(4, self.impt.ws.max_row + 1)]
        self.assertEqual(test_values, values)

    def test_get_max_site_id(self):
        value = self.impt._get_max_site_id()
        test_value = ['VS', '101578']
        self.assertEqual(test_value, value)

    def test_get_max_event_id(self):
        value = self.impt._get_max_event_id()
        test_value = ['VE', '17566']
        self.assertEqual(test_value, value)

    def test_add_ids_writes_correctly(self):
        test_values = []
        values = []
        self.impt._add_ids()
        file = openpyxl.load_workbook('IMM_Template_with_ids.xlsx')
        correct_file = openpyxl.load_workbook('IMM import template_with_ids_correct.xlsx')
        results = file['IMM_template']
        actual_results = correct_file['IMM_template']
        keys = [actual_results.cell(row=3, column=i).value for i in range(1, actual_results.max_column + 1)]
        for row in range(4, actual_results.max_row + 1):
            value_dict = {keys[i - 1]: actual_results.cell(row=row, column=i).value
                          for i in range(1, actual_results.max_column + 1)}
            test_values.append(value_dict)
        for row in range(4, results.max_row + 1):
            value_dict = {keys[i - 1]: results.cell(row=row, column=i).value
                          for i in range(1, actual_results.max_column + 1)}
            values.append(value_dict)
        self.assertEqual(test_values, values)

    def test_import_collection_events(self):
        self.fail("Not Implemented")

    def test_import_geographic_sites(self):
        self.impt._to_test()
        self.impt._import_site()
        query = 'Select * from GeographicSite' +\
                'where geo_site_id = (select max(geo_site_id) from GeographicSite'
        values = self.impt.cursor.execute(query).fetchone()
        test_values = []
        self.assertEqual(test_values, values)

    def test_import_taxonomy(self):
        self.fail("Not Implemented")

    def test_import_specimen(self):
        self.impt._import_specimen()

    def test_write_to_test(self):
        self.impt._to_test()
        value = self.impt._connection.getinfo(pyodbc.SQL_DATA_SOURCE_NAME)
        test_value = 'ImportTest'
        self.assertEqual(test_value, value)

    def test_write_to_prod(self):
        self.impt._to_prod()
        value = self.impt._connection.getinfo(pyodbc.SQL_DATA_SOURCE_NAME)
        test_value = 'IMM Prod'
        self.assertEqual(test_value, value)

    def test_generate_event(self):
        value = self.impt._generate_events()
        test_values = {'VE17567':
                       {'Bait': None,
                        'Collection method': 'Hand',
                        'Date': datetime(2018, 12, 31, 0, 0),
                        'Date remarks': None,
                        'Discipline': 'INV',
                        'Event Number': 'VE17567',
                        'Field Event Code': 'EV1',
                        'Net/Gear/Trap type': 'trap',
                        'Note': None,
                        'Permit Number': None,
                        'Season': 'Winter',
                        'Start time': None,
                        'Stop time': None,
                        'Time Standard': None,
                        'Trapping/Sampling Duration': None,
                        'Vessel Name': None,
                        'Collector': ['Hugh MacIntosh', 'Evan Harley'],
                        'Air temperature': None,
                        'Air temperature unit': None,
                        'Cloud cover': None,
                        'Weather remarks': None,
                        'Wind direction': None,
                        'Wind speed': None,
                        'Wind speed unit': None,
                        },
                       'VE17568':
                       {'Bait': None,
                        'Collection method': 'Hand',
                        'Date': datetime(2018, 12, 31, 0, 0),
                        'Date remarks': None,
                        'Discipline': 'INV',
                        'Event Number': 'VE17568',
                        'Field Event Code': 'EV2',
                        'Net/Gear/Trap type': 'trap',
                        'Note': None,
                        'Permit Number': None,
                        'Season': 'Winter',
                        'Start time': None,
                        'Stop time': None,
                        'Time Standard': None,
                        'Trapping/Sampling Duration': None,
                        'Vessel Name': None,
                        'Collector': ['Meg Sugrue', 'David Stewart'],
                        'Air temperature': None,
                        'Air temperature unit': None,
                        'Cloud cover': None,
                        'Weather remarks': None,
                        'Wind direction': None,
                        'Wind speed': None,
                        'Wind speed unit': None,
                        }}
        self.assertEqual(test_values, value)

    def test_write_spreadsheet(self):
        test_values = {
            'sheet_names': ('IMM_template', 'Person', 'Taxon', 'Site', 'Event'),
            'row_count': {'IMM_template': 5,
                          'Person': 7,
                          'Taxon': 3,
                          'Site': 3,
                          'Event': 3}
            }
        values = {}
        self.impt.write_spreadsheet()
        file = openpyxl.load_workbook(self.impt.data_filename)

        values['sheet_names'] = set(file.sheetnames)
        values['row_count'] = {}
        for sheet in file.sheetnames:
            worksheet = file[sheet]
            values['row_count'][sheet] = worksheet.max_row

        self.assertEqual(test_values, values)

    def test_check_sheets(self):
        self.assertTrue(self.impt._check_sheets())

    def test_check_persontaxa(self):
        test_values = [0, 'Complete']
        status, message = self.impt._check_persontaxa()
        self.assertEqual(test_values, [status, message])

    def test_check_collector(self):
        self.impt._to_test()
        test_values = [0, 1]
        values = []
        for thing in (340287, 340291):
            seq_num = self.impt._check_collector(thing)
            values.append(seq_num)
        self.assertEqual(test_values, values)

    def test_get_item_id(self):
        test_value = 2056010
        value = self.impt._get_item_id('V209394')
        self.assertEqual(test_value, value)
        return 0

    def test_write_update(self):
        self.impt.max_id = 200000
        test_result = self.impt._write_item_query(4, True)
        result = "Update Item\nset status_cd = 'catalog', area_cd = '', catalogue_num = 'TEST_CN_01', description = 'Test item 1'"+\
            "\nwhere item_id = 200000"
        self.assertEqual(test_result, result)
        print(result[1].format(*result[2]))

    def test_write_insert(self):
        self.impt.max_id = 2000000
        test_result = self.impt._write_item_query(4)
        result = "Insert into Item (item_id, status_cd, area_cd, catalogue_num, description)\n"+\
           "VALUES (2000000, 'catalog', '', 'TEST_CN_01', 'Test item 1')" 
                
        self.assertEqual(test_result, result)


    def test_get_person_org_colnames(self):
        test_vals = self.impt._get_person_org_colnames()
        vals = {}
        self.assertDictEqual()

    def test_prep_persons(self):
        row = self.impt.ws[4]
        test_vals = self.impt._prep_persons(row, 1)
        vals = {}
        self.assertEqual(test_vals, vals)


if __name__ == '__main__':
    unittest.main()
