import os
import pickle
import sqlalchemy
from sqlalchemy import exc
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.exceptions import InvalidFileException
import pandas
from pubsub import pub
from datetime import datetime
import urllib

'''Tools for the Bulk Import of Royal BC Museum Specimens'''


class ImportTools:
    '''Tool for importing data into Integrated Museum Management. Royal BC Museum's main collection management tool'''
    def __init__(self, *args, **kwargs):
        
        self.data_filename = ''
        self.discipline = ''
        self.area_cd = ''
        conn_str = "DRIVER={ODBC Driver 17 for SQL Server};SERVER=RBCMIMMLIVE;DATABASE=Mastodon;UID=appuser;PWD=Museum2019**;"
        self._connection_string = urllib.parse.quote_plus(conn_str)
        self._engine = None
        self.data_file = None
        self.ws = None
        self.keys = None
        self.max_id = self._query_item_id()
        self.max_col = None
        self.write_status = {'ArchaeologicalSite': False,
                             'ArchaeologicalCollectionEvent': False,
                             'GeographicSite': False, 
                             'CollectionEvent': False,
                             'Taxonomy': False,
                             'Triggers': False}
        self.proc_log = []

    def _get_file(self, filename):
        '''Takes file name, opens the excel file as an pandas dataframe object then sets up some
        other necessary attributes'''
        if not os.path.exists('files\\'):
            os.mkdir('files')
        self.data_filename = filename
        head, tail = os.path.split(self.data_filename)
        try:
            self.data_file = openpyxl.Workbook("IMM_template")
            if os.path.exists(f'files\\{tail[:-5]}_keys.pkl'):
                self.ws = pandas.read_excel(filename, sheet_name="IMM_template", engine="openpyxl")
            else:
                self.ws = pandas.read_excel(filename, sheet_name="IMM_template", header=1, engine="openpyxl")
        except InvalidFileException as e:
            return -1, "Did you cancel? Invalid File Error"
        self._get_prog_info()
        self.max_col = self._set_max_col()
        self._set_keys()
        try:
            self.ws['Date'] = self.ws.Date.astype('datetime64')
            self.ws['Date Identified'] = self.ws['Date Identified'].astype('datetime64')
            self.ws['Catalogue Number'] = self.ws['Catalogue Number'].astype('str')
        except AttributeError as e:
            print(e)
            return -1, "Key Not Found in Worksheet, it is likely that the Keys file doesn't match the current Spreadsheet"
        return 0, None

    def _set_max_col(self):
        '''Gets the number of columns in the spreadsheet'''
        row = list(self.ws)
        col_count = max([i for i in range(len(row))if row[i] is not None])
        return col_count      

    def _set_keys(self, reload=False, add_ids={}):
        '''gets the spreadsheets column headers for the purpose of having keys for the dicionaries used later'''

        head, tail = os.path.split(self.data_filename)
        if os.path.exists(f'files\\{tail[:-5]}_keys.pkl') and not reload:
            self.keys = pickle.load(open(f'files\\{tail[:-5]}_keys.pkl', 'rb'))
        elif reload:
            self.keys.update(add_ids)
            filename = f'{tail[:-5]}_keys.pkl'
            pickle.dump(self.keys, open(f'files\\{filename}', 'wb'))
        else:
            key_list = [key for key in list(self.ws) if not key.startswith('Unnamed')]
            if self.discipline:
                self.keys = {key: (self.ws[key][0] if not self.ws[key][0].startswith('[DISCIPLINE]') else 
                               self.ws[key][0].replace('[DISCIPLINE]', f'{self._get_full_disc()}Item')) for key in key_list}
            else:
                self.keys = {key: self.ws[key][0] for key in key_list}
            self.ws.drop(0, inplace=True)
            filename = f'{tail[:-5]}_keys.pkl'
            pickle.dump(self.keys, open(f'files\\{filename}', 'wb'))

    def _write_prog(self):
        ''' Updates the progress log for the file loaded '''
        prog_log = open('files\\prog_log.log', 'a')
        prog_log.write(f'{self.data_filename}: {self.proc_log[-1]}\n')
        return 0

    def _get_prog_info(self):
        # Reads the progress log for the file loaded
        temp = []
        try:
            prog_log = open('files\\prog_log.log', 'r')
        except FileNotFoundError:
            self.proc_log = ['New Import']
            return 0
        for row in prog_log:
            if row.startswith(self.data_filename):
                temp = row[row.find(': ') + 1:].strip(' ').split(', ')
        if temp == []:
            temp = ['New Import']
        self.proc_log = [value.strip() for value in temp]
        return 0

    def _find_person_organization(self, type="Person"):
        ''' Return all unique persons in the spreadsheet for import
         Persons to be a dict in format {personName: [ids]}'''
        max = self.ws.shape[0]
        pub.sendMessage('UpdateMessage', message=f'Finding Unique {type}', update_count=2, new_max=max)
        person_cols = self._find_relevant_column(type)
        names = []
        persons = {}
        for col in person_cols:
            values = self.unique_non_null(self.ws[col])
            translation = str.maketrans(";:|-", ",,,,")
            values_trans = [name.translate(translation) for name in values]
            values_split = []
            for name in values_trans:
                values_split.extend(name.split(','))
            values_split = [name.strip() for name in values_split]
            
            names.extend(values_split)
        names = list(set(names))

        for i in range(len(names)):
            if ',' in names[i]:
                name = [thing.strip() for thing in names[i].split(',')]
                names[i] = ' '.join(name)
        names = list(set(names))
        mess = f'Querying the database for {type}s'
        pub.sendMessage('UpdateMessage', message=mess, update_count=2, new_max=2*len(names))
        with self._engine.connect() as connection:
            for name in names: 
                mess = f'Querying the database for {name}'
                pub.sendMessage('UpdateMessage', message=mess)
                if "'" in name:
                    name = name.replace("'", "''")
                query = self._find_person_query(name, type)
                try:
                    results = connection.execute(query).fetchall()
                except:
                    print(query)
                    print('pause')

                persons[name] = []
                if results != []:
                    for i in range(len(results)):
                        persons[name].append(results[i][0])
                else:
                    persons[name] = ['NEW?']
                mess = f'{name} not found in database' if persons[name] == ['NEW?'] \
                    else f'{name} found. Person_id: {persons[name]}'
                pub.sendMessage('UpdateMessage', message=mess)
        return persons

    def unique_non_null(self, series):
        '''Helper function which returns the unique non-null values from the DataFrame'''
        return list(series.dropna().unique())

    def _find_person_query(self, name, type = 'Person' ):
        '''Return the id for various person types including Person, Artist, and Maker'''
        table = type
        column = f'{type}_id'
        if type == "Person":
            query = f"select {column} from {table} where search_name = '{name}'"
        else:
            query = f"select {column} from {table} where org_name = '{name}'"
        return query


    def _get_full_disc(self):
        '''Returns the whole disicpline name'''
        disciplines = {
            'bot': 'Botany',
            'ent': 'Entomology',
            'geo': 'Geology',
            'her': 'Herpetology',
            'ich': 'Ichthyology',
            'inv': 'Invertebrate',
            'mam': 'Mammalogy',
            'orn': 'Ornithology',
            'pal': 'Paleontology',
            'history':'Modern History',
            'archeolg':'Archaeology',
            'ethnolg':'Ethnology',
            }
        return disciplines[self.discipline]

    def _find_relevant_column(self, method):
        ''' Return the index of the columns relevant to the _find methods above.
        Values in a list of indices '''
        relevant_cols = []
        disc = self._get_full_disc()
        table_ids = {'Person': ['Person.search_name'],
                     'Organization': ['Organization.org_name'],
                     'Taxon': ['Taxon.term'],
                     'Events': ['CollectionEvent.', 'ArchaeologicalCollectionEvent.'],
                     'Sites': ['GeographicSite.', 'GeoSiteNote'],
                     'SitesImpt': ['GeographicSite.'],
                     'Item': ['Item'],
                     'NHItem': ['NaturalHistoryItem.', 'GeographicSite.collector', 'CollectionEvent.event_num'],
                     'FieldMeasurement': ['FieldMeasurement.'],
                     'DisciplineItem': ['[DISCIPLINE].', disc + 'Item.'],
                     'ImptTaxon': ['Taxonomy.', 'taxon_id'],
                     'Preparation': ['Preparation.'],
                     'ChemicalTreatment': ['ChemicalTreatment.'],
                     'Maker': ['MakerOrganization.org_name'],
                     'Artist': ['Artist.search_name'],
                     'HHItem':['HumanHistoryItem.'],
                     'EthItem': ['EthnologyItem.'],
                     'ArcItem': ['ArchaeologyItem.'],
                     'ArcSite': ['ArchaeologicalSite.'],
                     'Technique': ['Technique.'],
                     'Material': ['Material.'],
                     'MHist': ['ModernHistoryItem.'],
                     'Location': ['Location.location_cd'],
                     'ImptLocation':['Location.location_id'],
                     'OtherNumber': ['OtherNumber.'],
                     'GeoSiteNote':['GeographicSite.geo_site_id', 'GeoSiteNote']}
        table_id = table_ids[method]

        if len(table_id) > 1:
            for key, value in self.keys.items():
                if any([value.startswith(id) for id in table_id]):
                       relevant_cols.append(key)
                
        else:
            id = table_id[0]
            for key, value in self.keys.items():
                if value.startswith(id):
                    relevant_cols.append(key)

        return list(set(relevant_cols))
        
    def _split_persons(self, person_names):
        ''' Returns the split value of person names where a delineator is present'''
        delineators = ";:|/\\"
        person_names = person_names.replace(',', '~')
        if any(char in person_names for char in delineators):
            for char in delineators:
                if char in person_names:
                    person_names = person_names.replace(char, ',')
            names = [name.strip(' ') for name in person_names.split(',')]
        else:
            names = person_names
        if isinstance(names, list):
            names = [name.replace('~', ',') for name in names]
        else:
            names = names.replace('~', ',')
        return names

    def _find_taxa(self):
        ''' Returns the unique taxa for each taxon in the import spreadsheet
        Taxa to be a dict in format {scientificName: [ids]}'''
        max = len(self.ws['Scientific Name'].unique())*2
        pub.sendMessage('UpdateMessage', message=f'Finding unique Taxa', update_count=2, new_max=max)
        taxon_cols = self._find_relevant_column('Taxon')
        taxa = {} 
        sns = self.unique_non_null(self.ws[taxon_cols[0]])
        mess = 'Querying database for Taxa'
        pub.sendMessage('UpdateMessage', message=mess, update_count=2, new_max=2*len(sns))
        for sn in sns:
            mess = 'Querying database for {sn}'
            pub.sendMessage('UpdateMessage', message=mess)
            taxa[sn] = []
            taxa[sn].extend(self._query_taxa(sn))
            mess = f'{sn} not found in database' if taxa[sn] == ['NEW?'] \
                else f'{sn} found. Taxon_id: {taxa[sn]}'
            pub.sendMessage('UpdateMessage', message=mess)

        return taxa

    def _query_taxa(self, sn):
        '''Returns the taxon ids of a scientific name'''
        with self._engine.connect() as connection:
            if not sn.endswith('sp.') or sn.find(' ') == -1:
                query = f"Select * from ScientificName where scientific_name ='{sn}' and discipline_cd = '{self.discipline}'"
            else:
                sn = sn[: sn.find(' ')]
                query = f"Select taxon_id, term from taxon where term = '{sn}' and discipline_cd = '{self.discipline}'"
            results = connection.execute(query).fetchall()
            if results != []:
                taxa = [result[0] for result in results]
            else:
                taxa = ["NEW?"]
            
        return taxa

    def _find_locations(self):
        '''Finds unique location codes then queries the database for the location_id'''
        max = self.ws.max_row
        pub.sendMessage('UpdateMessage', message=f'Finding unique Locations', update_count=2, new_max=max)
        loc_col = self._find_relevant_column('Location')[0]
        loc = {}
        loc_cds = list(self.ws[loc_col].unique())
        mess = 'Querying database for Locations'
        pub.sendMessage('UpdateMessage', message=mess, update_count=2, new_max=2*len(loc_cds))
        for cd in loc_cds:
            loc[cd] = self._query_loc_id(cd)
            mess = f'{cd} not found in database' if loc[cd] == ['NEW?'] \
                else f'{cd} found. Location_id: {loc[cd]}'
            pub.sendMessage('UpdateMessage', message=mess)
        return loc

    def _query_loc_id(self, loc_cd):
        ''' Queries the database for location code passed into the method.'''
        with self._engine.connect() as connection:
            query = f"Select location_id from Location where location_code ='{loc_cd}'"
            result = connection.execute(query).fetchone()
            if result != [] and result is not None:
                return [result[0]]
            else:
                return ['NEW?']

    def _generate_sites(self):
        ''' Generates new sites for import, from unique sites in the import spreasheet'''
        area_dict = {'natural': "GeographicSite.collector_site_id",
                     'human': "ArchaeologicalSite.temporary_num"}
        site_type = area_dict[self.area_cd]
        new_site_id = self._get_max_site_id()
        relevant_cols = self._find_relevant_column('Sites')
        working_sheet = self.ws.copy()
        working_sheet = working_sheet[relevant_cols]
        cols_with_variation = {col: working_sheet[col].nunique() for col in relevant_cols if working_sheet[col].nunique() > 0}
        generated_sites = working_sheet[working_sheet.duplicated(subset = cols_with_variation, keep='first')]
        site_col = next((key for key, value in self.keys.items() if value == site_type))
        site_numbers = [f'{new_site_id[0]}{str(int(new_site_id[1])+i)}' for i in range(len(generated_sites))]
        generated_sites[site_col] = site_numbers
        return generated_sites

    def _get_max_site_id(self):
        ''' Queries the database for the highest collector_site_id for the discipline selected
        returns the discipline specific prefix and the site_id'''
        with self._engine.connect() as connection:
            prefix_query = f"Select geo_site_prefix from NHDisciplineType where discipline_cd = '{self.discipline}'"
            prefix = connection.execute(prefix_query).fetchall()[0][0]
            query = "Select max(convert(int, SUBSTRING(collector_site_id, 3, 100))) from GeographicSite " + \
                f"where discipline_cd = '{self.discipline}' and substring(collector_site_id, 1, 2) = '{prefix}'"
            result = connection.execute(query).fetchone()
            if len(str(result)) > 6:
                diff = 6 - len(str(result))
                prefix += ''.join({'0' for i in range(0, diff)})
            max_site_id = [prefix, str(result[0])]
        return max_site_id

    def _get_max_event_id(self):
        '''Queries the database for the highest event_num for the discipline selected
        returns the discipline specific prefix and the event_num'''
        with self._engine.connect() as connection:
            if self.area_cd == 'natural':
                prefix_query = "Select coll_event_prefix from NHDisciplineType where discipline_cd = " +\
                                    f"'{self.discipline}'"
                prefix = connection.execute(prefix_query).fetchall()[0][0]
                query = "Select max(convert(int, SUBSTRING(event_num, 3, 100))) from CollectionEvent " + \
                f"where discipline_cd = '{self.discipline}' and substring(event_num, 1, 2) = '{prefix}'"
            else:
                prefix = 'CE'
                query = "Select max(convert(int, SUBSTRING(event_num, 3, 100))) from ArchaeologicalCollectionEvent "
            result = connection.execute(query).fetchone()
            if len(str(result)) > 6:
                diff = 6 - len(str(result))
                prefix += ''.join({'0' for i in range(0, diff)})
            max_event_id = [prefix, str(result[0])]
        return max_event_id

    def _generate_events(self):
        '''Generates new collection events for import, from the unique events in the import spreadsheet'''
        generated_events = {}
        new_event_id = self._get_max_event_id()
        relevant_cols = self._find_relevant_column('Events')
        working_sheet = self.ws.copy()
        working_sheet = working_sheet[relevant_cols]
        cols_with_variation = {col: working_sheet[col].nunique() for col in relevant_cols if working_sheet[col].nunique() > 0}
        generated_events = working_sheet[working_sheet.duplicated(subset = cols_with_variation, keep='first')]
        event_numbers = [f'{new_event_id[0]}{int(new_event_id[1])+i}' for i in range(len(generated_events))]
        generated_events['Event Number'] = event_numbers
        return generated_events

    def _write_persontaxa(self, data, section):
        ''' '''
        max = len(data)
        pub.sendMessage('UpdateMessage', message=f'Writing {section}', update_count=2, new_max=max)
        names = {'Person': ["Name", "person_id"], "Organization": ["Organization", "organization_id"], "Taxon": ["Term", "taxon_id"]}
        for key in data.keys():
            data[key] = ', '.join([str(data[key][i]) for i in range(len(data[key]))])
        data = {names[section][0]: list(data.keys()), names[section][1]: list(data.values())}
        data_to_write = pandas.DataFrame(data)
        for r in dataframe_to_rows(data_to_write, index=False, header=True):
            pub.sendMessage('UpdateMessage', message=f'Writing rows to spreadsheet')
            self.data_file[section].append(r)
        return 0

    def _write_siteevent(self, data, section):
        ''' '''
        max = len(data)
        pub.sendMessage('UpdateMessage', message=f'Writing {section}', update_count=2, new_max=max)
        # Writes sites and events to spreadsheet
        worksheet = self.data_file[section]
        for r in dataframe_to_rows(data, index=False, header=True):
            pub.sendMessage('UpdateMessage', message=f'Writing rows to spreadsheet')
            worksheet.append(r)
        return 0

    def _write_locations(self, data, tab):
        ''' '''
        pub.sendMessage('UpdateMessage', message=f'Writing Location', update_count=2, new_max=max)
        worksheet = self.data_file['Location']
        data = {'Location Code': list(data.keys()), 'location_id': list(data.values())}
        data_to_write = pandas.DataFrame(data)
        for r in dataframe_to_rows(data_to_write, index=False, header=True):
            worksheet.append(r)
        return 0


    def write_spreadsheet(self):
        '''Writes the found and generated data to new tabs in the import spreadsheet'''
        if self.area_cd == 'natural':
            missing = {'IMM_template', 'Person', 'Organization', 'Taxon', 'Site', 'Event'} - set(self.data_file.sheetnames)
            tabs = {'Person': [self._find_person_organization, self._write_persontaxa],
                    'Organization': [self._find_person_organization, self._write_persontaxa],
                    'Taxon': [self._find_taxa, self._write_persontaxa], 
                    'Site': [self._generate_sites, self._write_siteevent], 
                    'Event': [self._generate_events, self._write_siteevent]}
        elif self.discipline == 'history':
            missing = {'IMM_template', 'Location'} - set(self.data_file.sheetnames)
            tabs = {'Location' : [self._find_locations, self._write_locations]}
        else:
            missing = {'IMM_template', 'Person', 'Site', 'Event'} - set(self.data_file.sheetnames)
            tabs = {'Person': [self._find_person_organization, self._write_persontaxa],  
                    'Site': [self._generate_archsites, self._write_siteevent], 
                    'Event': [self._generate_events, self._write_siteevent]}
        if len(missing) > 0:
            for sheet in missing:
                self.data_file.create_sheet(sheet)
        pub.sendMessage('UpdateMessage', message='Writing Spreadsheet', update_count=1, new_max=4)

        for tab in tabs.keys():
            if tab in ('Person', 'Organization'):
                data = tabs[tab][0](type=tab)
            else:
                data = tabs[tab][0]()
            if isinstance(data, pandas.DataFrame) and data.empty:
                continue
            elif isinstance(data, dict) and data == {}:
                continue
            return_value = tabs[tab][1](data, tab)
            if return_value == 1:
                return 1
            
            pub.sendMessage('UpdateMessage', message=f"{tab}s Complete")

        for r in dataframe_to_rows(self.ws, index=False, header=True):
            self.data_file['IMM_template'].append(r)
        ws = self.data_file.get_sheet_by_name("Sheet")
        self.data_file.remove(ws)
        self.data_file.save(self.data_filename)
        self.proc_log.append('Write Spreadsheet')
        return 0 


    def _add_ids(self):
        '''Adds the relevant ids to the spreadsheet for later writing to the database'''
        self.data_file = openpyxl.load_workbook(self.data_filename)
        if self.area_cd == 'natural':
            expected = ['Person', 'Organization', 'Taxon']
        elif self.discipline == 'history':
            expected = ['Location']
        else:
            expected = ['Person']
        
        pub.sendMessage('UpdateMessage', message='Adding IDs to Spreadsheet', update_count=1, new_max=4)
        for sheet in expected:
            data = {sheet: {}}
            workingsheet = self.data_file[sheet]
            if sheet in ['Person', 'Organization', 'Taxon']:
                data[sheet] = {workingsheet.cell(i, 1).value: workingsheet.cell(i, 2).value
                               for i in range(2, workingsheet.max_row + 1)}
            else:
                keys = [workingsheet.cell(row=1, column=i).value for i in range(1, workingsheet.max_column + 1)]
                for row in range(2, workingsheet.max_row + 1):
                    id = workingsheet.cell(row=row, column=1).value
                    data[sheet][id] = {keys[i - 1]: workingsheet.cell(row=row, column=i).value
                                       for i in range(1, workingsheet.max_column + 1)}
            if sheet in ['Person', 'Organization', 'Taxon']:
                self._handle_persontaxa(data)
            else:
                self._handle_location(data)
            pub.sendMessage('UpdateMessage', message=f"{sheet} Complete")
        try:
            idx = self.data_file.sheetnames.index("IMM_template")
            ws = self.data_file.get_sheet_by_name("IMM_template")
            self.data_file.remove(ws)
            self.data_file.create_sheet("IMM_template", idx)
            for r in dataframe_to_rows(self.ws, index=False, header=True):
                self.data_file['IMM_template'].append(r)
            self.data_file.save(self.data_filename)
        except PermissionError as e:
            return e, 'Failed! The file is still open'
        self.proc_log.append('IDs added')
        return 0, 'Done'

    def _handle_persontaxa(self, data):
        '''Specific logic for writing person_ids and taxon_ids to the spreadsheet'''
        tab = list(data.keys())[0]
        data = data[tab]
        relevant_cols = sorted(self._find_relevant_column(tab))
        col_names = {'Collector': 'Collector.pid',
                     'Determinavit': 'Determinavit.pid',
                     'Preparator':'Preparator.pid',
                     'Collector Organization': 'Collector.oid',
                     'Determinavit Organization': 'Determinavit.oid',
                     'Preparator Organization':'Preparator.oid',
                     'Scientific Name': 'Taxonomy.taxon_id'}
        for col in relevant_cols:
            values = list(self.unique_non_null(self.ws[col]))
            if tab != 'Taxon':
                persons = [{'name': value, f'{col} ID': self._get_split_name_ids(value, data)} for value in values]
                persons = pandas.DataFrame.from_records(persons)
                if not persons.empty:
                    self.ws = pandas.merge(self.ws, persons, how='left', left_on=col, right_on="name")
                else:
                    continue
            else:
                taxa = [{'term': value, f'Taxon ID': data[value]} for value in values]
                taxa = pandas.DataFrame.from_records(taxa)
                self.ws = pandas.merge(self.ws, taxa, how='left', left_on=col, right_on='term')

            self._set_keys(reload=True, add_ids = {f'{col} ID': col_names[col]})

        return 0

    def _get_split_name_ids(self, name, data):
        '''Gets the ids from the person tab which match the names in the person column of the IMM Template tab'''
        names = self._split_persons(name)
        return ', '.join([data[names[i]] for i in range(len(names))])


    def _handle_location(self, data):
        '''Gets the locations ids from the Locations tab to match the data in the IMM Template tab'''
        data = data['Location']
        relevant_col = self._find_relevant_column('Location')
        locations = [{'location': value, f'Location ID': data[value]} for value in values]
        locations = pandas.DataFrame.from_records(locations)
        self.ws = pandas.merge(self.ws, locations, how='left', left_on=col, right_on='location')
        return 0


    def _to_prod(self):
        '''Sets the connection string and creates the connection to the Production Database'''
        self._engine = None
        conn_str = "Driver={ODBC Driver 17 for SQL Server};Server=RBCMIMMLIVE;Database=Mastodon;UID=appuser;PWD=Museum2019**;"
        self._connection_string = urllib.parse.quote_plus(conn_str)
        try:
            self._engine = sqlalchemy.create_engine("mssql+pyodbc:///?odbc_connect=%s" % self._connection_string)
        except sqlalchemy.exc.InterfaceError as e:
            print(e)
            return -1, 'Did you remember to re-assign user appuser when you last restored the database'
        self.max_id = self._query_item_id()
        return 0, "Database connection changed to Production"

    def _to_test(self):
        '''Sets the connection and creates the connection to the Test Database'''
        self._engine = None
        conn_str = "Driver={ODBC Driver 17 for SQL Server};Server=RBCMIMMSTAGING;Database=Import_Test;UID=appuser;PWD=Museum2019**;"
        self._connection_string = urllib.parse.quote_plus(conn_str)
        try:
            self._engine = sqlalchemy.create_engine("mssql+pyodbc:///?odbc_connect=%s" % self._connection_string)
        except sqlalchemy.exc.InterfaceError as e:
            print(e)
            return -1, 'Did you remember to re-assign user appuser when you last restored the database'
        self.max_id = self._query_item_id()
        return 0, 'Database connection changed to Test'

    def _import_site(self, connection):
        '''Logic for importing site data to the database
           pub methods are for updating the progress bar '''
        pub.sendMessage('UpdateMessage', message="Writing Sites",
                        update_count=2,
                        new_max=1)
        relevant_cols = self._find_relevant_column('SitesImpt')
        marker_col = "Collector's Site ID" if self.area_cd == 'natural' else 'Temporary Number'
        table = 'GeographicSite' if self.area_cd == 'natural' else 'ArchaeologicalSite'
        site_data = self.ws[relevant_cols].copy().drop_duplicates(marker_col, keep='first')
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in site_data.columns.tolist()}
        site_data.rename(columns=keys, inplace=True)
        site_data.dropna(1, "all", inplace=True)
        site_data['discipline_cd'] = self.discipline

        with connection.begin():
            try:
                site_data.to_sql(table, con=connection, if_exists='append', index=False)
            except ValueError as e:
                print(e)
                raise ValueError
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
            self.import_site_note(connection)
        pub.sendMessage('UpdateMessage', message="Sites Complete")
        return 0

    def import_site_note(self, connection):
        '''Logic for importing Geo Site Notes'''
        data = self.get_site_note_data(connection)
        relevant_cols = self._find_relevant_column('GeoSiteNote')
        relevant_cols.append('geo_site_id')
        data = data[relevant_cols]
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key !='geo_site_id'}
        data.rename(columns=keys, inplace=True)
        data.dropna(axis=0, how='all', subset=['note_date', 'title', 'note'], inplace=True)
        data.drop_duplicates(inplace=True)
        if data.empty:
          return 0  

        try:
            data.to_sql('GeoSiteNote', con=connection, if_exists='append', index=False)
        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError
        return 0


    def get_site_note_data(self, connection):
        '''Gets only the rows which have Collector Site IDs then gets the geo_site_ids from the database
           then merges the two datasets to create a dataset for import'''
        data = self.ws[self.ws["Collector's Site ID"].notnull()]
        site_ids = self._query_site_id('GeographicSite', connection)
        data = pandas.merge(data, site_ids, left_on="Collector's Site ID", right_on='collector_site_id', how = 'left')
        return data


    def _import_event(self, connection):
        '''Specific logic for importing event data into the database'''
        pub.sendMessage('UpdateMessage', message="Writing Events",
                        update_count=2,
                        new_max=1)
        relevant_cols = self._find_relevant_column('Events')
        table = 'CollectionEvent' if self.area_cd == 'natural' else 'ArchaeologicalCollectionEvent'
        marker_col = 'Event Number'
        data = self.ws[relevant_cols].copy().drop_duplicates(marker_col, keep='first')
        data['Date'] = data['Date'].dt.strftime('%Y-%m-%d')
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist()}
        data.rename(columns=keys, inplace=True)
        data.dropna(1, "all", inplace=True)
        data['discipline_cd'] = self.discipline
        try:
            data.to_sql(table, con=connection, if_exists='append', index=False)
        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError
        pub.sendMessage('UpdateMessage', message="Events Complete")

        return 0

    def _import_site_event(self, connection):
        '''Specific logic for importing the linkage between sites and events'''
        pub.sendMessage('UpdateMessage', message="Writing Site-Event Table",
                        update_count=2,
                        new_max=1)
        table = 'GeographicSite_CollectionEvent' if self.area_cd == 'natural' \
            else 'ArchaeologicalSite_Event'
        cols = ['geo_site_id', 'coll_event_id'] if self.area_cd == 'natural' \
            else ['site_id', 'event_id']
        site_table = 'GeographicSite' if self.area_cd == 'natural' \
            else 'ArchaeologicalSite'
        event_table = 'CollectionEvent' if self.area_cd == 'natural' \
            else 'ArchaeologicalCollectionEvent'
        sites = self._query_site_id(site_table, connection)
        events = self._query_event_id(event_table, connection)
        data_to_import = self.ws.copy()
        data_to_import = pandas.merge(data_to_import, sites, left_on="Collector's Site ID", right_on='collector_site_id', how='left')
        data_to_import = pandas.merge(data_to_import, events, left_on='Event Number', right_on='event_num', how='left')
        data_to_import = data_to_import.groupby(cols).agg({
                cols[0]: 'max',
                cols[1]: 'max'
            })
        data_to_import = data_to_import[cols]
        try:
            data_to_import.to_sql(table, con=connection, if_exists='append', index=False)

        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError
        pub.sendMessage('UpdateMessage', message="Site-Event Complete")
        return 0

    def _query_site_id(self, table, connection):
        '''Gets datbase ids for sites and events'''
        keys = {'GeographicSite': ['geo_site_id', 'collector_site_id'],
                'ArchaeologicalSite': ['site_id', 'temp_num']}
        sites = pandas.read_sql_query(f"select {keys[table][0]}, {keys[table][1]} from {table}", connection)
        return sites

    def _query_event_id(self, table, connection):
        '''Gets datbase ids for sites and events'''
        keys = {'CollectionEvent': ['coll_event_id', 'event_num'],
                'ArchaeologicalCollectionEvent': ['event_id', 'event_num']}
        events = pandas.read_sql_query(f"select {keys[table][0]}, {keys[table][1]} from {table}", connection)

        return events

    def _check_process(self, process):
        ''' Performs a check to verify that data to perform the process is present in the dataset'''
        processes = {'item': 'Item',
                         'nhitem': 'NHItem',
                         'disc_item': 'DisciplineItem',
                         'preparation': 'Preparation',
                         'taxonomy': 'ImptTaxon',
                         'persons': 'Person',
                         'ChemTreat': 'ChemicalTreatment',
                         'FieldMeas': 'FieldMeasurement',
                         'hhitem': 'HHItem',
                         'location': 'ImptLocation'}
        relevant_cols = self._find_relevant_column(processes[process])
        values = self.ws[relevant_cols].dropna(1, "all")
        if values.empty:
            return False
        return True


    def _import_specimen(self, connection, update=False):
        '''Main method for importing all data which hinges on the item_id
           This includes all of the item tables (item, naturalhistoryitem, [discipline]item)
           as well as taxonomy, preparation, collector, determinavit, etc. '''
        if self.area_cd == 'natural':
            processes = {'item': self._import_item,
                            'nhitem': self._import_nhitem,
                            'disc_item': self._import_discipline_item,
                            'preparation': self._import_preparation,
                            'taxonomy': self._import_taxon,
                            'ChemTreat': self._import_chem,
                            'FieldMeas': self._import_field_measurement,
                            'OtherNum': self._import_other_num
                        }
        elif self.discipline != 'history':
            processes = {'item':self._import_item,
                            'hhitem': self._import_hhitem,
                            'arcitem':self._import_arcitem,
                            'OtherNum': self._import_other_num
                            }
        else:
            processes = {'item': self._import_item,
                         'hhitem': self._import_hhitem,
                         'mhitem': self._import_mhitem,
                         'location': self._import_location}

        if not update:
            self.max_id += 1
        else:
            item_ids = self._get_item_id()
            self.ws = pandas.merge(self.ws, item_ids, on='Catalogue Number')
        if  self.area_cd == 'natural':
            to_do = ['item', 'nhitem', 'disc_item', 'preparation', 
                        'ChemTreat','taxonomy', 'FieldMeas'] 
        elif self.discipline != 'history':
            to_do = ['item', 'hhitem', 'arcitem']
        else:
            to_do = ['item', 'hhitem', 'mhitem', 'location']
        pub.sendMessage('UpdateMessage', message="Writing Specimens",
                        update_count=2,
                        new_max=len(processes.keys())+2)
        for process in to_do:
            if process != 'mhitem':
                if not self._check_process(process):
                    pub.sendMessage("UpdateMessage", message=f'{process} Complete')
                    continue
                else:
                    result = processes[process](connection, update)
            else:
                result = processes[process](update)
            pub.sendMessage("UpdateMessage", message=f'{process} Complete')
        pub.sendMessage('UpdateMessage', message="Importing Specimens Complete")
        return 0

    def write_persons_to_db(self, context='all'):
        '''The function called to import persons to the database'''
        pub.sendMessage('UpdateMessage', message="Writing Persons",
                        update_count=2,
                        new_max=1)
        pub.sendMessage('UpdateMessage', message='Setting Triggers to off',
                        update_count=2, new_max=1)
        self._set_triggers()
        with self._engine.connect() as connection:
            with connection.begin():
                result = self._import_person(connection)

        pub.sendMessage('UpdateMessage', message='Setting Triggers to on',
                        update_count=2, new_max=1)
        self._set_triggers()
        pub.sendMessage('UpdateMessage', message='Complete!!')
        self.proc_log.append('Import Complete')
        return 0


    def _query_item_id(self, disconnect=True):
        if self._engine is None:
            self._engine = sqlalchemy.create_engine("mssql+pyodbc:///?odbc_connect=%s" % self._connection_string)
        with self._engine.connect() as connection:
            max_query = "Select Max(item_id) from Item"
            max_id = connection.execute(max_query).fetchone()[0]

        return max_id + 1

    def _get_item_id(self):
        catalogue_nums = self.ws['Catalogue Number'].tolist()
        catalogue_nums_str = "', '".join(catalogue_nums)
        with self._engine.connect() as connection:
            query = f"select catalogue_num as 'Catalogue Number', item_id from Item where catalogue_num in ('{catalogue_nums_str}')"
            item_ids = pandas.read_sql_query(query, connection)
        return item_ids

    def _write_update(self, data, table):
        formats = {'item': 
                   {'table_name': 'Item', 
                    'format': f"set status_cd = 'catalog', area_cd = '{self.area_cd}', ",
                    'offset': 3},
                   'nhitem':
                   {'table_name':'NaturalHistoryItem', 
                    'format': f"set discipline_cd = '{self.discipline}', ",
                    'offset': 2},
                    'hhitem':
                    {'table_name':'HumanHistoryItem', 
                    'format': f"set discipline_cd = '{self.discipline}', ",
                    'offset': 2},
                   'discitem': {'table_name': f'{self._get_full_disc()}Item', 'format':"set ",
                   'offset': 1}}
        data = data.to_dict()
        keys = list(data.keys())
        keys.pop(keys.index('item_id'))
        string = ''
        for key in keys:
            if not string == '':
                string += f", {key} = '{data[key]}'"
            else:
                string = f"{formats[table]['format']} {key} = '{data[key]}'"

        query = f"Update {formats[table]['table_name']} {string} where item_id = {data['item_id']}"
        return query

    def _import_item(self, connection, update=False):
        relevant_cols = self._find_relevant_column('Item')
        relevant_cols.append('item_id')
        if not update:
            item_ids = [self.max_id+(i+1) for i in range(self.ws.shape[0])]
            self.ws['item_id'] = item_ids
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['status_cd'] = 'catalog'
            data_to_import['area_cd'] = self.area_cd
            try:
                data_to_import.to_sql('Item', con=connection, if_exists='append', index=False)
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
        else:
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['query'] = data_to_import.apply(lambda row: self._write_update(row, 'item'), axis=1)
            for query in data_to_import.query.to_list():
                try:
                    connection.execute(query)
                except exc.SQLAlchemyError as e:
                    print(e)
                    raise exc.SQLAlchemyError
        return 0

    def _import_hhitem(self, connection, update=False):
        relevant_cols = self._find_relevant_column('HHItem')
        relevant_cols.append("item_id")
        if not update:
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['discipline_cd'] = self.discipline
            try:
                data_to_import.to_sql('Item', con=connection, if_exists='append')
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
        else:
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['query'] = data_to_import.apply(lambda row: self._write_update(row, 'item'), axis=1)
            for query in data_to_import.query.to_list():
                try:
                    connection.execute(query)
                except exc.SQLAlchemyError as e:
                    print(e)
                    raise exc.SQLAlchemyError
        return 0

    def _import_mhitem(self, connection, update):
        if not update:
            data_to_import = self.ws['item_id']
            try:
                data_to_import.to_sql('ModernHistoryItem', con=connection, if_exists='append', index_label='item_id')
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
            return 0
        else:
            return 0

    def _write_arcitem_query(self, connection, update=False):
        relevant_cols = self._find_relevant_column('ArcItem')
        relevant_cols.append('item_id')
        if not update:
            data_to_import = self.ws[relevant_cols].copy()
            data_to_import.dropna(1, "all", inplace=True)
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            try:
                data_to_import.to_sql(f'ArchaeologyItem', con=connection, if_exists='append', index=False)
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
        else:
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['query'] = data_to_import.apply(lambda row: self._write_update(row, 'item'), axis=1)
            for query in data_to_import.query.to_list():
                try:
                    connection.execute(query)
                except exc.SQLAlchemyError as e:
                    print(e)
                    raise exc.SQLAlchemyError
        return 0

    def _import_nhitem(self, connection, update=False):
        if not update:
            sites = self._query_site_id('GeographicSite', connection)
            events = self._query_event_id('CollectionEvent', connection)
            relevant_cols = self._find_relevant_column('NHItem')
            relevant_cols.append('item_id')
            data_to_import = self.ws[relevant_cols].copy()
            data_to_import = pandas.merge(data_to_import, sites, left_on="Collector's Site ID", right_on='collector_site_id', how='left')
            data_to_import = pandas.merge(data_to_import, events, left_on='Event Number', right_on='event_num', how='left')
            data_to_import.drop(["Collector's Site ID", 'collector_site_id', 'Event Number', 'event_num'], axis=1, inplace=True)
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key not in ('item_id', 'geo_site_id', 'coll_event_id')}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['discipline_cd'] = self.discipline
            
            try:
                data_to_import.to_sql('NaturalHistoryItem', con=connection, if_exists='append', index=False)
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
        else:
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['query'] = data_to_import.apply(lambda row: self._write_update(row, 'item'), axis=1)
            for query in data_to_import.query.to_list():
                try:
                    connection.execute(query)
                except exc.SQLAlchemyError as e:
                    print(e)
                    raise exc.SQLAlchemyError
        return 0

    def _import_discipline_item(self, connection, update=False):
        relevant_cols = self._find_relevant_column('DisciplineItem')
        relevant_cols.append('item_id')
        if not update:
            data_to_import = self.ws[relevant_cols].copy()
            data_to_import.dropna(1, "all", inplace=True)
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            try:
                data_to_import.to_sql(f'{self._get_full_disc()}Item', con=connection, if_exists='append', index=False)
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
        else:
            data_to_import = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data_to_import.columns.tolist() if key != 'item_id'}
            data_to_import.rename(columns=keys, inplace=True)
            data_to_import.dropna(1, "all", inplace=True)
            data_to_import['query'] = data_to_import.apply(lambda row: self._write_update(row, 'item'), axis=1)
            for query in data_to_import.query.to_list():
                try:
                    connection.execute(query)
                except exc.SQLAlchemyError as e:
                    print(e)
                    raise exc.SQLAlchemyError
        return 0

    def _import_preparation(self, connection, update=False):
        relevant_cols = self._find_relevant_column('Preparation')
        relevant_cols.append('item_id')
        data = self.ws[relevant_col].copy()
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key != 'item_id'}
        data.rename(columns=keys, inplace=True)
        data.dropna(1, "all", inplace=True)
        data.dropna(0, "all", subset=[key for key in data.columns.tolist() if key != 'item_id'], inplace=True)
        try:
            data.to_sql(f'Preparation', con=connection, if_exists='append', index=False)
        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError

        return 0

    def _import_taxon(self, connection, update = False):
        if not update:
            relevant_cols = self._find_relevant_column('ImptTaxon')
            relevant_cols.append('item_id')
            data = self.ws[relevant_cols].copy()
            keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key != 'item_id'}
            data.dropna(0, "all", subset=[key for key in data.columns.tolist() if key != 'item_id'], inplace=True)
            data.dropna(1, "all", inplace=True)
            data.rename(columns=keys, inplace=True)
            data['accepted'] = 1
            data['cf'] = 0
            data['aff'] = 0
            try:
                data.to_sql(f'Taxonomy', con=connection, if_exists='append', index=False)
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError
        else:
            print('not implemented')
        return 0

    def _import_chem(self, connection, update=False):
        relevant_cols = self._find_relevant_column('ChemicalTreatment')
        relevant_cols.append('item_id')
        data = self.ws[relevant_col].copy()
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key != 'item_id'}
        data.dropna(0, "all", subset=[key for key in data.columns.tolist() if key != 'item_id'], inplace=True)
        data.dropna(1, "all", inplace=True)
        data.rename(columns=keys, inplace=True)
        data['seq_num'] = 0
        try:
            data.to_sql(f'ChemicalTreatment', con=connection, if_exists='append', index=False)
        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError
        return 0

    def _import_field_measurement(self, connection, update=False):
        relevant_cols = self._find_relevant_column('FieldMeasurement')
        relevant_cols.append('item_id')
        data = self.ws[relevant_cols].copy()
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key != 'item_id'}
        data.dropna(0, "all", subset=[key for key in data.columns.tolist() if key != 'item_id'], inplace=True)
        data.dropna(1, "all", inplace=True)
        data.rename(columns=keys, inplace=True)
        try:
            data.to_sql(f'FieldMeasurement', con=connection, if_exists='append', index=False)
        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError
        return 0

    def _import_location(self, connection, update=False):
        '''Imports location data into the database'''
        relevant_cols = self._find_relevant_column('ImptLocation')
        relevant_cols.append('item_id')
        data = self.ws[relevant_cols].copy()
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key != 'item_id'}
        data['seq_num'] = 0
        data.rename(columns=keys, inplace=True)
        if update:
            result = self._query_seq_num('ItemLocation')
            if result == False:
                return 1
        with self._engine.connect() as connection:
            data.to_sql(f'ItemLocation', con=connection, if_exists='append', index=False)
        return 0
        
    def _query_seq_num(self, table, connection):
        '''Given the table name and connection to the database gets the max sequence number in the table to match the item_ids'''
        item_ids = self.ws['item_id'].tolist()
        item_ids = "', '".join(catalogue_nums)
        with self._engine.connect() as connection:
            query = f"select item_id, max(seq_num)+1 from {table} where item_id in ('{item_ids}')"
            seq_nums = connection.execute(query).fetchall()
            seq_nums = [id[0] for id in item_ids]
        return seq_nums

    def _update_loc_status(self, vals, connection):
        ''' Helper Method for updating Locations - Sets the status of the items previous location to 0
            in preparation for the new location to be given the current location tag'''
        query = f"update itemlocation set status = 0 where item_id = {self.max_id} and seq_num = {vals['seq_num'] - 1}"
        try:
            self.cursor.execute(query)
        except:
            return False
        return True

    def _import_other_num(self, connection):
        '''Specific logic for importing OtherNumbers'''
        relevant_cols = self._find_relevant_column('OtherNumber')
        relevant_cols.append('item_id')
        data = self.ws[self.ws[relevant_cols]]
        keys = {key: self.keys[key][self.keys[key].find('.')+1:] for key in data.columns.tolist() if key != 'item_id'}
        data['seq_num'] = self._query_seq_num('OtherNumber')
        data.dropna(1, "all", inplace=True)
        data.rename(columns=keys, inplace=True)
        try:
            data.to_sql(f'OtherNumber', con=connection, if_exists='append', index=False)
        except exc.SQLAlchemyError as e:
            print(e)
            raise exc.SQLAlchemyError

        return 0
    
    def _query_taxonomy(self, connection):
        '''Queries the database for Taxonomy IDs'''
        query = "select item_id, max(taxonomy_id) as taxonomy_id from Taxonomy group by item_id"
        value = pandas.read_sql_query(query, connection)
        return value

    def _import_person(self, connection, context='only'):
        ''' Helper method for the specimen import method
           imports person data (collector, determinavit, preparator)'''
        types = {'Collector': 'Collector ID', 
                 'CollectorOrganization': 'Collector Organization ID', 
                 'Determinavit': 'Determinavit ID', 
                 'DeterminavitOrganization': 'Determinavit Organization ID', 
                 'Preparator': 'Preparator ID', 
                 'PreparatorOrganization': 'Preparator Organization ID'}
        for type in types.keys():
            if not types[type] in list(self.ws):
                continue
            relevant_cols = [types[type]]
            table = ''
            if type in ('Collector', 'CollectorOrganization'):
                relevant_cols.append('Event Number')
                table = 'Collector'
            else:
                if type in ('Determinavit', 'DeterminavitOrganization'):
                    table = 'Determinavit'
                else:
                    table = 'Preparator'
                relevant_cols.append('item_id')
            data = self.create_person_data(connection, table, relevant_cols)
            if data.empty:
                continue
            try:
                data.to_sql(table, con=connection, if_exists='append', index=False)
            except exc.SQLAlchemyError as e:
                print(e)
                raise exc.SQLAlchemyError 
        pub.sendMessage('UpdateMessage', message='Complete!!')
        self.proc_log.append('Import Persons')
        return 0

    def create_person_data(self, connection, table, cols):
        '''Helper method for creating the person data for import'''
        if cols[1] == 'item_id' and cols[1] not in list(self.ws):
            item_ids = self._get_item_id()
            self.ws = pandas.merge(self.ws, item_ids, open='Catalogue Number')
        data = self.ws[cols].drop_duplicates(keep='first').copy()
        data[cols[0]] = data[cols[0]].astype('string')
        data.dropna(subset=[cols[0]], inplace=True)
        if data.empty:
            return data
        
        if table == 'Collector':
            if data.empty:
                return data
            data = data.groupby(cols[1])[cols[0]].apply('; '.join).reset_index()
            events = self._query_event_id('CollectionEvent', connection)
            data = pandas.merge(data, events, left_on='Event Number', right_on='event_num', how='left')
            data_dict = data.to_dict(orient="records")
            person_data_dicts = []
            for record in data_dict:
                seq_num = 1
                event_num = record['coll_event_id']
                persons = record[cols[0]].split('; ')
                for person in persons:
                    person_data_dicts.append({'coll_event_id':event_num, 'collector_pid': person, 'seq_num': seq_num})
                    seq_num += 1
            person_data = pandas.DataFrame.from_records(person_data_dicts)
        elif table == 'Determinavit':
            
            taxonomy_ids = self._query_taxonomy(connection)
            data = pandas.merge(data, taxonomy_ids, on='item_id', how='left')
            data['seq_num'] = data.groupby(cols).cumcount()+1
            data.drop('item_id', axis=1, inplace=True)
            data_dict = data.to_dict(orient='records')
            person_data_dicts = []
            for record in data_dict:
                seq_num = 1
                taxonomy_id = record['taxonomy_id']
                persons = record[cols[0]].split('; ')
                for person in persons:
                    person_data_dicts.append({'taxonomy_id':taxonomy_id, 'determinavit_pid': person, 'seq_num': seq_num})
                    seq_num += 1
            person_data = pandas.DataFrame.from_records(person_data_dicts)
        else:
            data['seq_num'] = data.groupby(cols).cumcount()+1
            data.rename({cols[0]: self.keys[cols[0]][self.keys[cols[0]].find('.')+1:]}, axis=1,  inplace=True)
            person_data = data.copy()

        return person_data

    def _set_identity_insert(self, table, connection):
        ''' Allows/Disallows the inserting into the data tables'''
        if self.write_status[table] is False:
            query = f'set identity_insert {table} on;'
        else:
            query = f'set identity_insert {table} off;'
        connection.execute(query)
        self.write_status[table] = not self.write_status[table]
        return 0

    def _set_triggers(self):
        '''Enables/Disables all triggers on the database'''
        if self.write_status['Triggers'] is False:
            status = 'DISABLE'
        else:
            status = 'ENABLE'
        query = f'''{status} TRIGGER create_sname ON Taxon;
                   {status} Trigger set_is_component on Component;
                   {status} TRIGGER set_qualified_name ON Taxonomy;
                   {status} TRIGGER clear_item_name ON Taxonomy;
                   {status} TRIGGER update_person_search_name ON Person;
                   {status} TRIGGER update_artist_search_name ON Artist;
                   {status} TRIGGER create_sname on Taxon;
                   {status} TRIGGER create_location_code on Location;'''
        with self._engine.connect() as connection:
            connection.execute(query)
        self.write_status['Triggers'] = not self.write_status['Triggers']
        return 0

    def write_siteevent_to_db(self, context='only'):
        '''Method to write just site and event data to the database (step 1 for some disciplines)'''
        pub.sendMessage('UpdateMessage', message='Setting Triggers to off',
                        update_count=1, new_max=3)
        self._set_triggers()
        with self._engine.connect() as connection:
            with connection.begin():
                self._import_site(connection)
                self._import_event(connection)
                if not self.area_cd == 'human':
                    self._import_site_event(connection)
        pub.sendMessage('UpdateMessage', message='Setting Triggers to on')
        self._set_triggers()
        if context == 'only':
            pub.sendMessage('UpdateMessage', message='Complete!!')
        else:
            pub.sendMessage('UpdateMessage', message='Sites and Events Complete!', update_count=2, new_max=1)
        self.proc_log.append('Import GeographicSite and CollectionEvent')
        return 0

    def write_specimen_taxa_to_db(self, context='only', update=False):
        '''Method to write just specimen related data to the database (setp 2 for the above disciplines)'''
        pub.sendMessage('UpdateMessage', message='Setting Triggers to off',
                        update_count=2, new_max=1)
        self._set_triggers()
        with self._engine.connect() as connection:
            with connection.begin():
                self._import_specimen(connection, update)

        pub.sendMessage('UpdateMessage', message='Setting Triggers to on',
                        update_count=2, new_max=1)
        self._set_triggers()
        if self.area_cd == 'natural':
            raw_con = self._engine.raw_connection()
            raw_con.execute("exec BuildAllScientificNames @discipline_cd = '{self.discipline}'")
        if context =='only':
            pub.sendMessage('UpdateMessage', message='Complete!!')
        else:
            pub.sendMessage('UpdateMessage', message='Specimen and Taxa Complete!', update_count=2, new_max=1)
        self.proc_log.append('Import Specimens and Taxa Complete')
        return 0

    def write_to_db(self):
        '''Writes the data from the import spreadsheet to the database'''
        step_dict = {'history':[self.write_specimen_taxa_to_db],
                     'other': [self.write_siteevent_to_db, self.write_specimen_taxa_to_db, self.write_persons_to_db]}
        if self.discipline == 'history':
            steps = step_dict['history']
        else:
            steps = step_dict['other']
        for step in steps:
            step(context='all')
        pub.sendMessage('UpdateMessage', message='Complete!!')
        self.proc_log.append('Import Complete')
        return 0

    def update_db(self):
        '''Updates the data in the db to match data from the import spreadsheet'''
        pub.sendMessage('UpdateMessage', message='Setting Triggers to off',
                        update_count=2, new_max=1)
        self._set_triggers()
        with self._engine.connect() as connection:
            with connection.begin():
                self._import_specimen(connection, update=True)
        pub.sendMessage('UpdateMessage', message='Setting Triggers to on',
                        update_count=2, new_max=1)
        self._set_triggers()
        self.cursor.commit()
        pub.sendMessage('UpdateMessage', message='Complete!!')
        self.proc_log.append('Import Complete')
        return 0
